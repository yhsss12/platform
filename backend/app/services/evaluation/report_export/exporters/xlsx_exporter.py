from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from app.services.evaluation.report_export.export_options import ExportOptions
from app.services.evaluation.report_export.exporters._utils import (
    EPISODE_TABLE_HEADERS,
    VIDEO_TABLE_HEADERS,
    dash,
    episode_table_row,
    metric_display_value,
    metric_status_label,
    ordered_metric_rows,
    section_enabled,
    video_table_row,
)


def export_xlsx(data: dict[str, Any], out_dir: Path, options: ExportOptions) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("导出 Excel 需要安装 openpyxl") from exc

    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "Summary"

    if section_enabled(options, "basicInfo"):
        ws = wb["Summary"]
        ws.append(["字段", "值"])
        basic = data.get("basicInfo") or {}
        for key, label in [
            ("taskName", "任务名称"),
            ("jobId", "任务 ID"),
            ("evaluationTypeLabel", "评测类型"),
            ("associatedTaskName", "关联任务"),
            ("simulationPlatform", "仿真平台"),
            ("statusLabel", "状态"),
            ("evaluationObjectLabel", "评测对象"),
            ("modelAssetName", "模型资产"),
            ("robotType", "机器人"),
            ("createdAt", "创建时间"),
            ("finishedAt", "完成时间"),
        ]:
            ws.append([label, dash(basic.get(key))])
        if data.get("legacyNotice"):
            ws.append(["旧格式提示", data["legacyNotice"]])
    else:
        ws = wb["Summary"]
        ws.append(["提示", "未包含基础信息"])

    if section_enabled(options, "metricResults"):
        metrics_ws = wb.create_sheet("Metrics")
        metrics_ws.append(["指标", "值", "单位", "状态", "来源", "说明"])
        for entry in ordered_metric_rows(data):
            metrics_ws.append(
                [
                    dash(entry.get("displayName") or entry.get("metricId")),
                    metric_display_value(entry),
                    dash(entry.get("unit")),
                    metric_status_label(entry, include_reason=options.include_unavailable_metric_reasons),
                    dash(entry.get("source")),
                    dash(entry.get("reason") if not entry.get("available") else entry.get("description")),
                ]
            )

    if section_enabled(options, "episodeResults"):
        episodes_ws = wb.create_sheet("Episodes")
        episodes_ws.append(EPISODE_TABLE_HEADERS)
        for row in data.get("episodeResults") or []:
            episodes_ws.append(episode_table_row(row))

    run_ws = wb.create_sheet("RunMetrics")
    run_ws.append(["键", "值"])
    for key, value in sorted((data.get("runMetrics") or {}).items()):
        run_ws.append([key, dash(value)])

    if section_enabled(options, "videoInfo"):
        videos_ws = wb.create_sheet("Videos")
        videos_ws.append(VIDEO_TABLE_HEADERS)
        for row in data.get("videoInfo") or []:
            videos_ws.append(video_table_row(row))

    if section_enabled(options, "diagnostics"):
        diag_ws = wb.create_sheet("Diagnostics")
        diag = data.get("diagnostics") or {}
        diag_ws.append(["字段", "值"])
        for key, label in [
            ("summary", "摘要"),
            ("failureReason", "失败原因"),
            ("errorMessage", "异常信息"),
            ("timedOut", "是否超时"),
            ("abnormalExit", "进程异常退出"),
        ]:
            value = diag.get(key)
            if isinstance(value, bool):
                value = "是" if value else "否"
            diag_ws.append([label, dash(value)])
        diag_ws.append(["缺失文件", ", ".join(diag.get("missingFiles") or []) or "-"])
        diag_ws.append(["日志摘要", dash(diag.get("logTail"))])

    if section_enabled(options, "runtimeFiles"):
        files_ws = wb.create_sheet("RuntimeFiles")
        files_ws.append(["文件名", "相对路径", "是否存在", "大小", "说明"])
        for row in data.get("runtimeFiles") or []:
            files_ws.append(
                [
                    dash(row.get("filename")),
                    dash(row.get("relativePath")),
                    "是" if row.get("exists") else "否",
                    dash(row.get("sizeLabel")),
                    dash(row.get("description")),
                ]
            )

    raw_ws = wb.create_sheet("RawJSON")
    raw_ws.append(["reportJson"])
    raw_ws.append([json.dumps(data, ensure_ascii=False)[:32000]])

    path = out_dir / "report.xlsx"
    wb.save(path)
    return path
